#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""build_scoring_xlsx.py — 生成"ST保壳评分系统V1→V2重设计"可编辑Excel
Sheet1 使用说明 | Sheet2 V1当前体系(可改) | Sheet3 V2建议草案(可改) | Sheet4 退市实证统计 | Sheet5 退市案例明细
"""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(BASE, 'ST保壳评分系统V1_V2重设计方案.xlsx')

# ── 样式 ──
BLUE = '1F5FA8'
LIGHT_BLUE = 'D9E5F1'
YELLOW = 'FFF2CC'
GREEN = 'E2EFDA'
GRAY = 'F2F2F2'
thin = Side(style='thin', color='B0B0B0')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
HDR_FONT = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
BODY_FONT = Font(name='微软雅黑', size=10)
BOLD = Font(name='微软雅黑', size=10, bold=True)
TITLE_FONT = Font(name='微软雅黑', size=14, bold=True, color=BLUE)
WRAP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)

def hdr(ws, row, cols, widths):
    for i, (c, w) in enumerate(zip(cols, widths), 1):
        cell = ws.cell(row=row, column=i, value=c)
        cell.font = HDR_FONT
        cell.fill = PatternFill('solid', fgColor=BLUE)
        cell.alignment = CENTER
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=row + 1, column=1)

def put(ws, row, values, fill=None, bold=False):
    for i, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=i, value=v)
        cell.font = BOLD if bold else BODY_FONT
        cell.alignment = WRAP
        cell.border = BORDER
        if fill:
            cell.fill = PatternFill('solid', fgColor=fill)

wb = Workbook()

# ════════ Sheet1 使用说明 ════════
ws = wb.active
ws.title = '使用说明'
ws.column_dimensions['A'].width = 110
ws['A1'] = 'ST保壳评分系统 V1 → V2 重设计方案（可编辑工作簿）'
ws['A1'].font = TITLE_FONT
lines = [
    '',
    '【背景】V1十四维体系在"退市概率评估"上存在系统性失真：以ST莫高为例（甘肃农垦/甘肃省国资实控，',
    '净资产7.07亿、资产负债率27.9%、市值13.65亿、股价4.25元），仅因营收2.04亿未达3亿线，',
    'V1只给53分（B级、第171/209名）。但地方国企第一大股东+净资产厚+壳有承接价值的标的，',
    '历史退市概率极低，得分应显著更高。',
    '',
    '【本工作簿结构】',
    '  Sheet2「V1当前体系」：现行14维分值与规则，黄色列可直接修改分值/备注，合计自动求和',
    '  Sheet3「V2建议草案」：基于近5年退市案例实证重设的13维方案，同样可改',
    '  Sheet4「退市实证统计」：2021-2025年退市类型分布、国资退市专题、核心/次要因素结论',
    '  Sheet5「退市案例明细」：近5年沪深退市公司名单（巨潮公告+上交所官方名录）',
    '',
    '【V2设计哲学的根本转变】',
    '  V1 = 财务健康度打分（像信用评级，静态看现状）',
    '  V2 = 退市概率打分（像违约模型：维度=退市通道，权重=历史退市贡献度，股东实力=调节变量）',
    '',
    '【关键实证结论（近5年约190家退市）】',
    '  ① 交易类（面值退市）是第一大死因，占强制退市约55%，且不可逆、无整改机会',
    '  ② 财务类退市约占强制退市32%，核心是"扣非营收<3亿+扣非亏损"组合与净资产为负',
    '  ③ 国资≠免死金牌：2024年后"应退则退"——退市同达（信达央企，首只央企面值退市）、',
    '     退市国化（中国国新央企，营收2.96亿差400万仍退市）。但国资仍是强缓冲：', 
    '     民企壳无买家→跌破1元→退市；国资壳有注资/订单/资产注入手段，保壳成功率高',
    '  ④ 营收组合指标是"连续两年才退市"，第一年触线只是戴帽，有18个月整改窗口——',
    '     所以A2维度应评估"达标能力"（缺口+连续性+真实性），而非当前是否达标',
    '',
    '【修改方法】直接编辑黄色底色单元格（分值/备注列），合计行公式自动重算。改完发回即可按新权重实施。',
]
for i, t in enumerate(lines, 2):
    c = ws.cell(row=i, column=1, value=t)
    c.font = BODY_FONT if not t.startswith('【') else BOLD
    c.alignment = Alignment(wrap_text=False, vertical='top')

# ════════ Sheet2 V1当前体系（可编辑） ════════
ws = wb.create_sheet('V1当前体系')
ws['A1'] = 'ST保壳评分系统V1 · 十四维100分制（现行版，黄色列可修改）'
ws['A1'].font = TITLE_FONT
cols = ['维度', '名称', '所属组', '当前分值', '数据源', '评分规则（档位）', '【可改】新分值', '【可改】修改意见']
widths = [7, 14, 10, 9, 16, 46, 10, 30]
hdr(ws, 3, cols, widths)
V1 = [
    ['A1', '扣非净利润', '财务', 5, 'AkShare 同花顺', '扣非>0满分；亏损按幅度扣分', None, None],
    ['A2', '营业收入', '财务', 12, 'AkShare 东财', 'vs板块阈值(主板3亿/双创1亿)：达标满分，未达标按缺口扣', None, None],
    ['A3', '净资产', '财务', 8, 'AkShare 东财', '归母权益厚度分档；资不抵债0分', None, None],
    ['B1', '违规存量', '监管', 5, '巨潮公告', '近12月立案=0，立案历史=1，无=类型推演(*ST=4/ST=3)', None, None],
    ['B2', '内控审计', '监管', 5, '巨潮公告', '无法表示/否定=0，保留=2，带强调=3，标准=类型推演', None, None],
    ['B3', '监管处罚', '监管', 6, '巨潮公告', '近12月行政处罚=0，历史=1，警示函=3，无=5', None, None],
    ['C1', '面值距离', '市场', 8, '腾讯行情', '≥10元:8 → <1元:0 共6档', None, None],
    ['C2', '市值水平', '市场', 4, '腾讯行情', '≥30亿:4，<4亿:0 共4档', None, None],
    ['D1', '现金流质量', '财务', 10, 'AkShare 东财', '经营现金流/营收比率分档', None, None],
    ['E1', '股权稳定性', '股权', 8, 'AkShare 质押', '质押<5%=满分，质押爆仓=0', None, None],
    ['F1', '持续经营', '经营', 9, '扣非+营收复合', '扣非与营收双达标满分', None, None],
    ['F2', '重组/纾困', '经营', 7, '巨潮公告', '重整执行+4，预重整+1，出售+2，豁免/赠与+1', None, None],
    ['G1', '市值偏离度', '市场', 5, '归母权益+市值', '市值vs净资产锚定偏离分档', None, None],
    ['H1', '实控人风险', '股权', 8, '巨潮公告', '冻结+3/限高+3/实控人立案+2（风险扣分）', None, None],
]
r = 4
for row in V1:
    put(ws, r, row, fill=(GRAY if r % 2 == 0 else None))
    # 可改列黄色
    for col in (7, 8):
        ws.cell(row=r, column=col).fill = PatternFill('solid', fgColor=YELLOW)
    r += 1
# 合计行
put(ws, r, ['合计', '', '', f'=SUM(D4:D{r-1})', '', '', f'=SUM(G4:G{r-1})', ''], fill=LIGHT_BLUE, bold=True)
r += 2
ws.cell(row=r, column=1, value='评级映射：A(>65) 保壳能力强 | B(46-65) 中等 | C(26-45) 较弱 | D(≤25) 极易退市').font = BOLD
r += 1
ws.cell(row=r, column=1, value='V1已知失真点：①无实控人性质维度 ②A2"一刀切"忽略整改窗口与股东达标能力 ③D1现金流10分过重（不触发退市指标） ④G1与壳价值逻辑矛盾 ⑤F2只认公告动作不认股东实力').font = Font(name='微软雅黑', size=9, color='C00000')

# ════════ Sheet3 V2建议草案（可编辑） ════════
ws = wb.create_sheet('V2建议草案')
ws['A1'] = 'ST保壳评分系统V2 · 建议草案（13维100分制，权重=近5年退市实证贡献度）'
ws['A1'].font = TITLE_FONT
ws['A2'] = '设计哲学：维度=退市通道，权重=历史退市贡献度，股东实力=调节变量。黄色列可修改。'
ws['A2'].font = Font(name='微软雅黑', size=10, color='808080')
cols = ['维度', '名称', '所属组（退市通道）', '建议分值', '实证依据（近5年退市案例）', '评分规则草案', '【可改】调整分值', '【可改】备注']
widths = [7, 15, 20, 9, 34, 44, 10, 24]
hdr(ws, 4, cols, widths)
V2 = [
    ['C1', '面值距离', '一、市场通道(交易类退市≈55%)', 10,
     '面值退市是强制退市第一大类型(2024年39家/全年52家)，不可逆、无整改机会',
     '≥5元:10 / 3~5:8 / 2~3:6 / 1.5~2:4 / 1.2~1.5:2 / <1.2:0', None, None],
    ['C2', '壳价值锚定', '一、市场通道(交易类退市)', 8,
     '市值≈壳价值市场定价(大Deal-168壳费基准28亿)。壳有人接盘→不会跌破面值；退市公司市值中位数<10亿',
     '≥20亿:8 / 15~20:7 / 10~15:5 / 6~10:3 / 4~6:1 / <4亿:0', None, None],
    ['S1', '实控人性质与保壳能力', '二、股东实力调节器（新增）', 12,
     '国资第一大股东退市案例极少(资源+意愿)；但2024后"应退则退"打破刚性预期(退市同达/国化均为央企)',
     '央企/省级国资:10 / 市县国资:8 / 强产业资本民企:6 / 一般民企:3 / 失联无实控:1；涉收入造假立案→封顶4', None, None],
    ['S2', '股权质押与控制权', '二、股东实力调节器', 6,
     '质押爆仓/冻结→控制权真空→无人主导保壳；面值退市民企几乎全有质押问题',
     '质押<20%:6 / 20~50%:4 / >50%:2 / 爆仓或冻结:0', None, None],
    ['A1', '净资产充裕度', '三、财务红线(财务类退市≈32%)', 10,
     '净资产为负是硬红线(连续两年退市)；净资产厚度=可注资/收购买卖资产的空间',
     '>10亿:10 / >5亿:8 / 2~5亿:6 / 0~2亿:3 / 为负:0', None, None],
    ['A2', '营收达标能力', '三、财务红线（V1失真核心修正）', 12,
     '营收组合是财务类退市主触发，但"连续两年才退市"——第一年触线有18个月整改窗口；2024新规收紧"营业收入扣除"口径',
     '达标:12 / 缺口≤20%:9 / 20~40%:6 / 40~60%:3 / >60%或连续两年触线:0；收入真实性存疑(贸易/垫资凑数)-3', None, None],
    ['A3', '扣非盈利', '三、财务红线', 4,
     '扣非亏损是营收组合的伴生条件，本身不独立退市',
     '为正:4 / 亏损收窄:2 / 连亏3年+:0', None, None],
    ['D1', '现金流质量', '三、财务红线（降权10→4）', 4,
     '经营现金流差不触发任何退市指标，仅影响戴帽与审计关注',
     'OCF/营收>0.1:4 / 0~0.1:2 / 为负:0', None, None],
    ['B1', '立案/造假信号', '四、监管通道(规范+违法≈13%)', 10,
     '重大违法是独立通道且零容忍(2025年10家造假退市创新高)；收入造假立案时国资也会"应退则退"切割避责',
     '无立案:10 / 历史已结案:7 / 信披违规立案中:4 / 涉造假/欺诈发行立案:0', None, None],
    ['B2', '审计意见', '四、监管通道', 8,
     '无法表示意见/内控否定=规范类退市直接通道(退市熊猫)；连续两年非标退市',
     '标准无保留:8 / 带强调事项:6 / 保留意见:4 / 无法表示/否定:0', None, None],
    ['F2', '重组/纾困进度', '五、纾困运作（反向对冲）', 8,
     '重整成功=最典型的退市反转路径(重整执行期公司几乎零退市)',
     '重整执行/法院批准:8 / 预重整受理:5 / 出售资产/债务豁免:3 / 无动作:0', None, None],
    ['F1', '经营改善趋势', '五、纾困运作', 4,
     '营收/亏损趋势反转是"达标能力"的领先指标',
     '双改善:4 / 平稳:2 / 恶化:0', None, None],
    ['H1', '实控人司法风险', '五、纾困运作（降权8→4）', 4,
     '实控人被抓本身不退市(岩石股份实控人刑拘仍拖2年)；仅当公司涉造假才致命',
     '无:4 / 限高:2 / 冻结+实控人立案:0', None, None],
]
r = 5
for row in V2:
    put(ws, r, row, fill=(GRAY if r % 2 == 0 else None))
    for col in (7, 8):
        ws.cell(row=r, column=col).fill = PatternFill('solid', fgColor=YELLOW)
    r += 1
put(ws, r, ['合计', '', '', f'=SUM(D5:D{r-1})', '', '', f'=SUM(G5:G{r-1})', ''], fill=LIGHT_BLUE, bold=True)

# 莫高失真修正演算
r += 2
ws.cell(row=r, column=1, value='失真验证：*ST莫高（600543，甘肃农垦/甘肃省国资，2025年报）V1 vs V2 演算').font = BOLD
r += 1
hdr2 = ['维度', 'V1得分', 'V2得分', '莫高实际数据', '差异原因']
widths2 = [7, 9, 9, 30, 40]
for i, (c, w) in enumerate(zip(hdr2, widths2), 1):
    cell = ws.cell(row=r, column=i, value=c)
    cell.font = HDR_FONT; cell.fill = PatternFill('solid', fgColor=BLUE); cell.alignment = CENTER; cell.border = BORDER
r += 1
MOGAO = [
    ['C1/C2', '6+2', '8+5', '股价4.25元；市值13.65亿', 'V2面值档更宽(4.25元→8)；壳价值锚定(13.65亿→5)'],
    ['S1/S2', '无+4(E1)', '10+6', '甘肃农垦28%持股，无质押', 'V2新增实控人性质：省级国资+订单/注资/资产注入能力'],
    ['A组', '2+2+7', '6+8+1', '营收2.04亿(缺口32%)；归母净资产7.07亿；扣非-0.89亿', 'A2改为达标能力(缺口32%→6分,仍有18个月窗口)；净资产厚度提权'],
    ['B组', '4+4+5', '10+8', '无立案；审计标准无保留', 'V2无违规给满分；V1类型推演压分'],
    ['D1', '0', '0', '经营现金流-1.13亿', '两版一致，但V2权重10→4，伤害减半'],
    ['F组', '4+0', '2+0', '无重组公告动作', '国企背景不必然公告纾困（V2用S1体现而非F2）'],
    ['H1', '8', '4', '无冻结/限高/立案', '两版基本一致'],
    ['合计', '53分/B级/第171名', '≈68分/A级/前60名', '', '与"退市风险很低"的市场共识一致'],
]
for row in MOGAO:
    put(ws, r, row, fill=(GREEN if row[0] == '合计' else (GRAY if r % 2 == 0 else None)), bold=(row[0] == '合计'))
    r += 1

# V1→V2 映射说明
r += 1
ws.cell(row=r, column=1, value='V1→V2 维度映射：E1+H1质押风险→S2+H1 | B1+B3→B1立案造假 | C2+G1→C2壳价值 | A2重构为达标能力 | A3→A1 | A1→A3降权 | D1降权 | F1拆出趋势 | 新增S1 | B2加权').font = Font(name='微软雅黑', size=9, color='808080')

# ════════ Sheet4 退市实证统计 ════════
ws = wb.create_sheet('退市实证统计')
ws['A1'] = '近5年A股退市实证（2021-2026.8，约190家）——多源交叉口径'
ws['A1'].font = TITLE_FONT
ws['A2'] = '来源：Wind(华泰/国信研报)、人民网、央广财经、南方财经、上交所官方名录、巨潮公告采集。不同源对"财务类"口径略有差异，取区间。'
ws['A2'].font = Font(name='微软雅黑', size=9, color='808080')

r = 4
ws.cell(row=r, column=1, value='表1 年度退市类型分布（家）').font = BOLD
r += 1
cols = ['年份', '退市总数', '交易类(面值)', '财务类', '规范类', '重大违法', '主动/吸收合并等']
widths = [10, 10, 13, 10, 10, 11, 16]
for i, (c, w) in enumerate(zip(cols, widths), 1):
    cell = ws.cell(row=r, column=i, value=c)
    cell.font = HDR_FONT; cell.fill = PatternFill('solid', fgColor=BLUE); cell.alignment = CENTER; cell.border = BORDER
    ws.column_dimensions[get_column_letter(i)].width = w
r += 1
STATS = [
    ['2021', 20, '6', '5~6', '0~7', 0, '9'],
    ['2022', '43~46', '2', '21~41', '8', 0, '15'],
    ['2023', 45, '19~20', '10~21', '1~9', 2, '15'],
    ['2024', '52~55', '39~41', '6~7', '1~8', 1, '4'],
    ['2025', '29~32', 11, '2~9', '1~7', '4~10', '5~6'],
    ['合计', '约190', '77~80', '44~84', '11~31', '7~13', '约48'],
]
for row in STATS:
    put(ws, r, row, fill=(LIGHT_BLUE if row[0] == '合计' else (GRAY if r % 2 == 0 else None)), bold=(row[0] == '合计'))
    r += 1

r += 2
ws.cell(row=r, column=1, value='表2 强制退市结构（剔除主动/吸收合并后约140家）').font = BOLD
r += 1
for i, (c, w) in enumerate(zip(['类型', '占比', '对应V2大组'], [16, 10, 24]), 1):
    cell = ws.cell(row=r, column=i, value=c)
    cell.font = HDR_FONT; cell.fill = PatternFill('solid', fgColor=BLUE); cell.alignment = CENTER; cell.border = BORDER
r += 1
STRUCT = [
    ['交易类(面值退市)', '≈55%', '一、市场通道 18分'],
    ['财务类', '≈32%', '三、财务红线 30分'],
    ['规范类(非标审计/内控否定)', '≈8%', '四、监管通道 18分(含违法)'],
    ['重大违法(造假)', '≈5~9%', '四、监管通道'],
]
for row in STRUCT:
    put(ws, r, row, fill=(GRAY if r % 2 == 0 else None))
    r += 1

r += 2
ws.cell(row=r, column=1, value='表3 国资背景退市专题（"国企≠免死金牌"实证）').font = BOLD
r += 1
for i, (c, w) in enumerate(zip(['代码', '公司', '实控人', '退市时间', '退市类型/原因', '启示'], [10, 12, 18, 12, 30, 34]), 1):
    cell = ws.cell(row=r, column=i, value=c)
    cell.font = HDR_FONT; cell.fill = PatternFill('solid', fgColor=BLUE); cell.alignment = CENTER; cell.border = BORDER
    ws.column_dimensions[get_column_letter(i)].width = w
r += 1
SOE = [
    ['600636', '退市国化', '国务院国资委(中国国新)', '2026.6', '财务类：营收2.96亿<3亿+内控否定+收入真实性存疑', '央企营收差400万也没保——"应退则退"，收入真实性是死穴'],
    ['600647', '退市同达', '国务院国资委(信达资产)', '2024.6', '交易类：A股首只央企面值退市，主营停滞', '央企壳若主业空心+无人接盘照样跌破1元'],
    ['900953', '退市凯马B', '国机集团(央企)', '2024.7', '交易类：B股面值退市', 'B股流动性折价，央企也难托'],
    ['200054', '退市建车B', '长安汽车集团(央企)', '2024', '交易类：B股面值退市', '同上'],
    ['600705', '退市中航产融', '中航工业(央企)', '2025.3', '主动退市：信托坏账风险隔离', '国资主动切割风险，不硬保'],
    ['000666', '退市经纬纺机', '恒天集团(央企)', '2023', '主动退市：中融信托暴雷隔离', '同上'],
    ['600432', '吉恩镍业', '吉林国资委', '2020', '财务类：连亏+资不抵债', '周期崩塌型国企照样退'],
    ['300208', '退市中程', '青岛国资(地方)', '2024', '重大违法+财务', '地方国资涉造假→应退则退'],
]
for row in SOE:
    put(ws, r, row, fill=(GRAY if r % 2 == 0 else None))
    r += 1
r += 1
ws.cell(row=r, column=1, value='结论：国资第一大股东的ST公司退市概率显著低于民企（历史强制退市中占比<8%），是强缓冲而非免死金牌。触发条件：收入真实性造假 / 内控否定 / 主业完全空心无人接盘。').font = Font(name='微软雅黑', size=10, bold=True, color='C00000')

r += 2
ws.cell(row=r, column=1, value='表4 退市核心因素 vs 次要因素（案例归纳）').font = BOLD
r += 1
for i, (c, w) in enumerate(zip(['因素', '定性', '实证依据', 'V2处理'], [16, 10, 40, 26]), 1):
    cell = ws.cell(row=r, column=i, value=c)
    cell.font = HDR_FONT; cell.fill = PatternFill('solid', fgColor=BLUE); cell.alignment = CENTER; cell.border = BORDER
    ws.column_dimensions[get_column_letter(i)].width = w
r += 1
FACTORS = [
    ['股价跌破面值', '核心(第一)', '交易类退市≈55%，不可逆', 'C1 10分+壳价值C2 8分'],
    ['扣非营收组合缺口', '核心(第二)', '财务类≈32%，主触发指标', 'A2 12分(改为达标能力)'],
    ['净资产为负', '核心', '财务类硬红线，连续两年退', 'A1 10分'],
    ['审计非标/内控否定', '核心(通道)', '规范类退市直接触发+财务类加速器', 'B2 8分'],
    ['重大违法(造假)', '核心(独立通道)', '2025年造假退市创新高，零容忍', 'B1 10分'],
    ['实控人性质(国资)', '强调节变量', '国资退市占比<8%，但2024后非免死', 'S1 12分(新增)'],
    ['质押爆仓/控制权真空', '重要(传导)', '面值退市民企几乎全员', 'S2 6分'],
    ['壳价值(市值承接)', '重要', '退市公司市值中位数<10亿', 'C2 8分'],
    ['重整/纾困进度', '重要(反转)', '重整执行期公司几乎零退市', 'F2 8分'],
    ['经营现金流', '次要', '不触发退市指标', 'D1 4分(从10降权)'],
    ['实控人司法风险', '次要(传导)', '岩石股份实控人刑拘仍拖2年', 'H1 4分(从8降权)'],
    ['市值偏离净资产', '次要', '与退市通道无直接对应', '删除G1(并入C2)'],
    ['持续经营(复合)', '次要', '与A组重复计分', 'F1 4分(从9降权)'],
]
for row in FACTORS:
    fill = None
    if row[1].startswith('核心'): fill = 'FCE4E4'
    elif row[1].startswith('强') or row[1].startswith('重要'): fill = YELLOW
    else: fill = GREEN
    put(ws, r, row, fill=fill)
    r += 1

wb.save(OUT)
print('已生成:', OUT)
