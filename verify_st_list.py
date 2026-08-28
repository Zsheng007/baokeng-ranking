#!/usr/bin/env python3
"""verify_st_list.py — 东财ST名单 × 沪深交易所官方名录 交叉核验

数据源（均为交易所官方接口）：
  - 上交所：query.sse.com.cn commonQuery.do（全A股公司名录，含科创板/主板/B股）
  - 深交所：www.szse.cn/api/report/ShowReport（A股/B股列表 xlsx，列结构两表一致：
            A股代码=列4 A股简称=列5 B股代码=列9 B股简称=列10）

比对逻辑：
  1. 从交易所官方名录构建两个集合：
     - official_all: 全部在市代码（DELIST_DATE='-'，供"是否退市"判断）
     - official_st : 在市且简称含ST（正则排除英文名误报）
  2. 与东财ST板块名单（st_names.json）按代码比对
  3. 输出差异：
     [1] 官方ST但东财名单缺失 → 疑似东财板块滞后（新戴帽/沪B股缺口）
     [2a] 东财有、官方在市但简称无ST → 再用腾讯行情简称仲裁：
          腾讯也无ST → 东财板块残留（疑似已摘帽，应剔除）
          腾讯有ST   → 官方名录名称滞后（保留，人工核实）
     [2b] 东财有、官方名录已退市/查无此码 → 严重：东财名单含已退市标的
     [3] 简称不一致（仅提示）
  4. 北交所标的（92/83/87开头）不参与沪深比对，单独列出

坑（历史教训）：
  - 上交所 COMPANY_ABBR 的 ST 前缀不齐全（*ST禾信带前缀，600165显示"宁科生物"），
    但腾讯行情证实600165/600525确实已不带ST → 是东财板块残留（已摘帽），
    故"东财有/官方无ST"必须腾讯仲裁，不能直接判官方名称滞后
  - 深交所英文名称含 "ST" 子串（Distillery/RealEstate），须用前后无英文字母的正则
  - 东财板块不含沪市B股ST（900915 ST中路系统性漏掉）
  - 上交所 B_STOCK_CODE='-' 表示无B股，不是有效代码

运行：用 venv python（需 requests + openpyxl）
  C:/Users/xiaot/.workbuddy/binaries/python/envs/default/Scripts/python.exe verify_st_list.py
"""

import json
import random
import re
import time
from datetime import datetime
from io import BytesIO

import requests
from openpyxl import load_workbook

SSE_URL = 'http://query.sse.com.cn/commonQuery.do'
SZSE_URL = 'https://www.szse.cn/api/report/ShowReport'
SSE_HEADERS = {
    'Referer': 'http://www.sse.com.cn/',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)',
}
SZSE_HEADERS = {
    'Referer': 'https://www.szse.cn/',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)',
}

BJ_PREFIX = ('92', '83', '87')  # 北交所代码前缀

# ST标记正则：*ST / SST / ST，前后不得是英文字母（排除 Distillery/RealEstate）
ST_RE = re.compile(r'(?<![A-Za-z])\*?S?ST(?![A-Za-z])')


def norm(name):
    """规范化简称：去空格（含全角）、全角＊转半角*"""
    if not name:
        return ''
    s = re.sub(r'[\s\u3000]+', '', str(name))
    return s.replace('＊', '*')


def is_st_name(name):
    return bool(ST_RE.search(norm(name)))


def fetch_sse():
    """上交所官方名录 → (official_all, official_st) 均为 {code: name}"""
    params = {
        'sqlId': 'COMMON_SSE_CP_GPJCTPZ_GPLB_GP_L',
        'PRODUCT_TYPE': '1',
        'pageHelp.pageSize': '3000',
    }
    r = requests.get(SSE_URL, params=params, headers=SSE_HEADERS, timeout=30)
    r.raise_for_status()
    rows = r.json()['result']
    all_map, st_map = {}, {}
    for row in rows:
        if row.get('DELIST_DATE') not in ('-', '', None):
            continue  # 已退市
        name = norm(row.get('COMPANY_ABBR', ''))
        for key in ('A_STOCK_CODE', 'B_STOCK_CODE'):
            code = (row.get(key) or '').strip()
            if code and code != '-':  # '-' 表示无该类股票
                all_map[code] = name
                if is_st_name(name):
                    st_map[code] = name
    return all_map, st_map, len(rows)


def fetch_szse():
    """深交所官方A股/B股列表 → (official_all, official_st)"""
    # 两张表列结构一致：A股代码4/A股简称5/B股代码9/B股简称10
    A_CODE, A_NAME, B_CODE, B_NAME = 4, 5, 9, 10
    all_map, st_map = {}, {}
    total = 0
    for tabkey in ('tab1', 'tab2'):
        params = {
            'SHOWTYPE': 'xlsx',
            'CATALOGID': '1110',
            'TABKEY': tabkey,
            'random': f'{random.random():.16f}',
        }
        r = requests.get(SZSE_URL, params=params, headers=SZSE_HEADERS, timeout=60)
        r.raise_for_status()
        wb = load_workbook(BytesIO(r.content))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        total += len(rows) - 1
        for row in rows[1:]:
            if not row or len(row) <= B_NAME:
                continue
            for ci, ni in ((A_CODE, A_NAME), (B_CODE, B_NAME)):
                code = str(row[ci] or '').strip()
                name = norm(row[ni])
                if code:
                    all_map[code] = name
                    if is_st_name(name):
                        st_map[code] = name
        time.sleep(1.0)
    return all_map, st_map, total


def tencent_names(codes):
    """腾讯行情批量查简称（GBK解码）：{code: name}，用于第三方仲裁"""
    if not codes:
        return {}
    def prefix(c):
        if c.startswith('6') or c.startswith('9'):
            return 'sh' + c
        return 'sz' + c
    out = {}
    # 每次最多50只，避免URL过长
    for i in range(0, len(codes), 50):
        batch = codes[i:i + 50]
        url = 'http://qt.gtimg.cn/q=' + ','.join(prefix(c) for c in batch)
        r = requests.get(url, timeout=20)
        raw = r.content.decode('gbk', errors='ignore')
        for line in raw.split(';'):
            line = line.strip()
            if '=' not in line:
                continue
            f = line.split('=')[1].strip('"').split('~')
            if len(f) > 2 and f[2]:
                out[f[2]] = f[1]
        time.sleep(0.5)
    return out


def main():
    print('=' * 60)
    print('ST名单交叉核验：东财ST板块 × 沪深交易所官方名录')
    print('=' * 60)

    sse_all, sse_st, sse_total = fetch_sse()
    print(f'\n上交所官方名录 {sse_total} 行 → 在市 {len(sse_all)} 只，其中名称带ST {len(sse_st)} 只')
    szse_all, szse_st, szse_total = fetch_szse()
    print(f'深交所官方列表 {szse_total} 行 → 在市 {len(szse_all)} 只，其中名称带ST {len(szse_st)} 只')

    official_all = {**sse_all, **szse_all}
    official_st = {**sse_st, **szse_st}

    # 东财名单
    ours = json.load(open('st_names.json', encoding='utf-8'))
    print(f'\n东财ST板块名单: {len(ours)} 家')

    ours_hs = {c: n for c, n in ours.items() if not c.startswith(BJ_PREFIX)}
    ours_bj = {c: n for c, n in ours.items() if c.startswith(BJ_PREFIX)}

    # 交叉比对（限沪深）
    miss = {c: official_st[c] for c in official_st if c not in ours_hs}   # 官方ST/东财无
    # 东财有/官方非ST → 先按在市状态分，再用腾讯简称仲裁
    suspects = {c: n for c, n in ours_hs.items() if c not in official_st and c in official_all}
    extra_delisted = {c: n for c, n in ours_hs.items() if c not in official_all}
    tencent = tencent_names(list(suspects.keys()))
    eastmoney_lag = {}    # 腾讯也无ST → 东财板块残留（疑似已摘帽，应剔除）
    official_name_lag = {}  # 腾讯有ST → 官方名录名称滞后（保留）
    for c, n in suspects.items():
        tn = tencent.get(c, '')
        if is_st_name(tn):
            official_name_lag[c] = {'ours': n, 'official': official_all[c], 'tencent': tn}
        else:
            eastmoney_lag[c] = {'ours': n, 'official': official_all[c], 'tencent': tn}
    name_diff = {c: {'ours': ours_hs[c], 'official': official_st[c]}
                 for c in ours_hs if c in official_st and norm(ours_hs[c]) != norm(official_st[c])}

    # 报告
    print('\n' + '-' * 60)
    print(f'[1] 交易所官方ST但东财名单缺失: {len(miss)} 只')
    for c, n in sorted(miss.items()):
        print(f'    {c} {n}  ← 疑似东财板块滞后，需核实是否新戴帽')
    print(f'[2a] 东财有、官方在市但简称无ST → 腾讯仲裁为东财板块残留（疑似已摘帽，应剔除）: {len(eastmoney_lag)} 只')
    for c, v in sorted(eastmoney_lag.items()):
        print(f'    {c} 东财[{v["ours"]}] 官方[{v["official"]}] 腾讯[{v["tencent"]}]')
    print(f'[2b] 东财有、官方在市、腾讯仍有ST → 官方名录名称滞后（保留榜单，人工核实）: {len(official_name_lag)} 只')
    for c, v in sorted(official_name_lag.items()):
        print(f'    {c} 东财[{v["ours"]}] 官方[{v["official"]}] 腾讯[{v["tencent"]}]')
    print(f'[2c] 东财有、官方已退市/查无此码: {len(extra_delisted)} 只（严重：榜单疑似含已退市标的）')
    for c, n in sorted(extra_delisted.items()):
        print(f'    {c} {n}')
    print(f'[3] 简称不一致（仅提示，不影响代码比对）: {len(name_diff)} 只')
    for c, v in sorted(name_diff.items()):
        print(f'    {c} 东财[{v["ours"]}] vs 官方[{v["official"]}]')
    print(f'[4] 北交所标的 {len(ours_bj)} 只（不参与沪深官方比对）: '
          + ', '.join(f'{c}{n}' for c, n in sorted(ours_bj.items())))

    hard_fail = bool(miss or extra_delisted or eastmoney_lag)
    verdict = 'FAIL' if hard_fail else ('DIFF' if (official_name_lag or name_diff) else 'PASS')
    print('\n' + '=' * 60)
    print(f'结论: {verdict} | 官方沪深在市 {len(official_all)} 只，官方ST {len(official_st)} '
          f'| 东财沪深 {len(ours_hs)} | 漏 {len(miss)} | 东财残留 {len(eastmoney_lag)} '
          f'| 官方名称滞后 {len(official_name_lag)} | 疑似退市残留 {len(extra_delisted)}')

    # 落盘
    report = {
        'meta': {
            'verified_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'verdict': verdict,
            'sse_scanned': sse_total,
            'szse_scanned': szse_total,
            'official_hs_total': len(official_all),
            'official_hs_st': len(official_st),
            'eastmoney_hs': len(ours_hs),
            'bj_count': len(ours_bj),
        },
        'missing_in_eastmoney': miss,
        'eastmoney_lag_unst': eastmoney_lag,
        'official_name_lag': official_name_lag,
        'delisted_in_eastmoney': extra_delisted,
        'name_mismatch': name_diff,
        'bj_stocks': ours_bj,
    }
    with open('st_list_verify.json', 'w', encoding='utf-8') as f:
        json.dump(report, f, ensure_ascii=False, indent=1)
    print('报告已写入 st_list_verify.json')


if __name__ == '__main__':
    main()
