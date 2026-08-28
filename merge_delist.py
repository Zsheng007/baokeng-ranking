#!/usr/bin/env python3
"""merge_delist.py — 合并SSE官方名录+巨潮采集+规则知识，构建最终退市案例表并写入Excel Sheet5"""
import json, os
from openpyxl import load_workbook
from datetime import datetime

BASE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(BASE, 'ST保壳评分系统V1_V2重设计方案.xlsx')

# ── 已知案例类型知识库（高置信） ──
KNOWN = {
    # 交易类(面值)
    '600466': 'trade',   # 蓝光发展 面值
    '601258': 'trade',   # 庞大集团 面值
    '603555': 'trade',   # 贵人鸟 面值
    '600823': 'trade',   # 世茂 面值
    '600277': 'trade',   # 亿利 面值
    '600565': 'trade',   # 迪马 面值
    '600297': 'trade',   # 广汇汽车 面值
    '600213': 'trade',   # 亚星 面值
    '600220': 'trade',   # 阳光 面值
    '600112': 'trade',   # 天成 面值
    '600836': 'trade',   # 易连 面值
    '600077': 'trade',   # 宋都 面值
    '600290': 'trade',   # 华仪 面值
    '600139': 'trade',   # 西源 面值
    '600311': 'trade',   # 荣华 面值
    '600393': 'trade',   # 粤泰 面值
    '600647': 'trade',   # 同达 央企面值
    '600532': 'trade',   # 未来 面值
    '600781': 'trade',   # 辅仁 面值
    '600242': 'trade',   # 中昌 面值
    '600767': 'trade',   # 运盛 面值
    '600209': 'trade',   # 罗顿 面值
    '600275': 'trade',   # 昌鱼 面值
    '600856': 'trade',   # 中天 面值
    '600091': 'trade',   # 明科 面值
    '600652': 'trade',   # 游久 面值
    '600890': 'trade',   # 中房 面值
    '600870': 'trade',   # 厦华 面值
    '600145': 'trade',   # 新亿 面值
    '603157': 'trade',   # 拉夏 面值
    '603996': 'trade',   # 中新 面值
    '600291': 'trade',   # 西水 面值
    '600695': 'trade',   # 绿庭 面值
    '600146': 'trade',   # 环球 面值
    '600385': 'trade',   # 金泰 面值
    '600555': 'trade',   # 海创 面值
    # 重大违法/造假
    '600260': 'fraud',   # 凯乐 造假
    '603603': 'fraud',   # 博天 造假
    '688086': 'fraud',   # 紫晶 欺诈发行
    '688555': 'fraud',   # 泽达 欺诈发行
    '600122': 'fraud',   # 宏图 造假(地方国资)
    '600093': 'fraud',   # 易见 造假
    '600090': 'fraud',   # 济堂 造假
    '600225': 'fraud',   # 卓朗 造假
    '600636': 'financial', # 国化 央企财务类(营收2.96亿+内控否定)
    # 规范类
    '600599': 'compliance',  # 熊猫 无法表示意见+内控否定
    '600696': 'financial',   # 岩石 营收0.39亿+保留+内控否定
    '600421': 'financial',   # 华嵘 扣非营收1.46亿 连续两年
    '600083': 'trade',    # 博信 面值
    '600306': 'trade',    # 商城 面值
    '600321': 'trade',    # 正源 面值
    '600766': 'trade',    # 园城 面值
    '603133': 'trade',    # 碳元 面值
    '600647': 'trade',
    # 财务类
    '600898': 'financial',  # 美讯
    '600070': 'financial',  # 富润
    '603963': 'trade',      # 大药 面值
    # 主动/吸收合并
    '600068': 'merger',   # 葛洲坝 吸收合并
    '600723': 'merger',   # 首商 吸收合并
    '600837': 'merger',   # 海通证券 吸收合并
    '600705': 'voluntary', # 中航产融 主动
}

def main():
    # 已知假阳性：可转债/债券"终止上市"误匹配（公司未退市）
    FALSE_POSITIVE = {'000735': '罗牛山(可转债终止上市误匹配)'}
    sse = json.load(open(os.path.join(BASE, 'sse_delisted_5y.json'), encoding='utf-8'))
    cn = json.load(open(os.path.join(BASE, 'delist_cases.json'), encoding='utf-8'))['cases']
    cn = {k: v for k, v in cn.items() if k not in FALSE_POSITIVE}
    from collections import Counter

    merged = {}
    # 1) SSE官方（权威）
    for code, info in sse.items():
        merged[code] = {'name': info['name'], 'date': info['date'], 'src': 'SSE官方名录', 'cat': None}
    # 2) 巨潮补充（SZSE为主）
    for code, info in cn.items():
        if code in merged:
            merged[code]['src'] += '+巨潮'
        else:
            merged[code] = {'name': info['name'], 'date': info['date'], 'src': '巨潮公告', 'cat': None}
        # 巨潮分类回填
        if merged[code]['cat'] is None and info.get('category') not in (None, 'unknown'):
            merged[code]['cat'] = info['category']
    # 3) 知识库覆盖（最高置信）
    for code, cat in KNOWN.items():
        if code in merged:
            merged[code]['cat'] = cat
    # 4) 规则：2022年5-7月退市潮=面值退市潮（历史公认，置信中）
    for code, m in merged.items():
        if m['cat'] is None and (m['date'][:7] in ('2022-05', '2022-06', '2022-07')):
            m['cat'] = 'trade'
            m['src'] += '(退市潮规则)'
    # 5) 2024年4-10月批量退市潮以面值为主（置信中）
    for code, m in merged.items():
        if m['cat'] is None and m['date'][:4] == '2024':
            m['cat'] = 'trade*'
    # 6) 2025-2026 unknown标注
    for code, m in merged.items():
        if m['cat'] is None:
            m['cat'] = 'unknown'

    CAT_CN = {'trade': '交易类(面值)', 'trade*': '交易类(推断)',
              'financial': '财务类', 'compliance': '规范类', 'fraud': '重大违法',
              'voluntary': '主动退市', 'merger': '吸收合并', 'unknown': '未分类'}
    dist = Counter(m['cat'] for m in merged.values())
    print('合并总家数:', len(merged))
    print('类型分布:', {CAT_CN[k]: v for k, v in dist.most_common()})

    json.dump({'meta': {'merged_at': datetime.now().strftime('%Y-%m-%d %H:%M'),
                        'sources': ['SSE官方名录77家', '巨潮公告132家', '知识库覆盖', '2022退市潮规则']},
               'cases': merged},
              open(os.path.join(BASE, 'delist_cases_final.json'), 'w', encoding='utf-8'),
              ensure_ascii=False, indent=1)

    # ── 写入Excel Sheet5 ──
    wb = load_workbook(XLSX)
    ws = wb.create_sheet('退市案例明细')
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    BLUE = '1F5FA8'; GRAY = 'F2F2F2'; YELLOW = 'FFF2CC'; GREEN = 'E2EFDA'
    thin = Side(style='thin', color='B0B0B0')
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    HDR_FONT = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
    BODY = Font(name='微软雅黑', size=10)
    TITLE_FONT = Font(name='微软雅黑', size=14, bold=True, color=BLUE)

    ws['A1'] = '近5年沪深退市公司明细（2021.8-2026.8，合并口径）'
    ws['A1'].font = TITLE_FONT
    ws['A2'] = '来源：上交所官方名录(权威77家) + 巨潮终止上市公告采集(SZSE为主) + 知识库案例 + 退市潮规则推断。分类置信度：知识库>巨潮>规则推断。'
    ws['A2'].font = Font(name='微软雅黑', size=9, color='808080')
    cols = ['代码', '名称', '退市日期', '退市类型', '类型置信', '数据来源']
    widths = [10, 14, 12, 14, 10, 26]
    for i, (c, w) in enumerate(zip(cols, widths), 1):
        cell = ws.cell(row=4, column=i, value=c)
        cell.font = HDR_FONT; cell.fill = PatternFill('solid', fgColor=BLUE)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A5'
    r = 5
    for code, m in sorted(merged.items(), key=lambda x: x[1]['date'], reverse=True):
        conf = '高' if '知识' not in m['src'] and '*' not in m['cat'] and m['cat'] != 'unknown' else (
               '中' if ('*' in m['cat'] or '规则' in m['src']) else '低')
        row = [code, m['name'], m['date'], CAT_CN[m['cat']], conf, m['src']]
        for i, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=i, value=v)
            cell.font = BODY; cell.border = BORDER
            cell.alignment = Alignment(horizontal=('left' if i in (2, 6) else 'center'), vertical='center')
            if r % 2 == 0:
                cell.fill = PatternFill('solid', fgColor=GRAY)
        r += 1
    # 统计尾行
    r += 1
    ws.cell(row=r, column=1, value='合计').font = Font(name='微软雅黑', size=10, bold=True)
    ws.cell(row=r, column=2, value=f'{len(merged)} 家').font = Font(name='微软雅黑', size=10, bold=True)
    ws.cell(row=r, column=4, value='；'.join(f'{CAT_CN[k]}:{v}' for k, v in dist.most_common())).font = Font(name='微软雅黑', size=10, bold=True)
    wb.save(XLSX)
    print('Sheet5 已写入, 总行数:', r)

if __name__ == '__main__':
    main()
